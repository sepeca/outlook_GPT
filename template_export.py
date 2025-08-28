import shutil
from typing import List, Dict, Any, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill

def _detect_header_row(ws: Worksheet, max_cols: int = 120, search_rows: int = 60) -> Tuple[int, List[str]]:
    """Detect a 2-row header: row1 = group, row2 = leaf; combine if both exist."""
    # r = 2  # old: force row 2 only
    r1, r2 = 1, 2  # new

    vals1 = [ws.cell(row=r1, column=c).value for c in range(1, max_cols + 1)]
    vals2 = [ws.cell(row=r2, column=c).value for c in range(1, max_cols + 1)]

    # trim right tail by last non-empty in row2
    while vals2 and (vals2[-1] in (None, "")):
        vals2.pop()
        vals1.pop()

    labels = []
    for a, b in zip(vals1, vals2):
        a = ("" if a in (None, "") else str(a).strip())
        b = ("" if b in (None, "") else str(b).strip())
        if b and a:
            labels.append(f"{a} | {b}")  # combined
        elif b:
            labels.append(b)             # leaf only
        else:
            labels.append(a)             # fallback to group text if leaf empty

    return r2, labels


def _unmerge_data_area(ws: Worksheet, first_data_row: int = 3) -> None:
    """Unmerge any merged ranges that touch the data area (>= first_data_row)."""
    to_unmerge = [str(mr) for mr in ws.merged_cells.ranges if mr.max_row >= first_data_row]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def export_rows_to_template(
    template_path: str,
    out_path: str,
    sheet_name: str,
    rows: List[Dict[str, Any]],
    start_row: int | None = None,
) -> None:
    """
    Copy Excel template and append `rows` under its header.
    - Header labels are read from the template.
    - If start_row is None -> write right under header_row.
    - Maps JSON keys to human headers via aliases, for example "Název Klienta".
    """
    shutil.copyfile(template_path, out_path)
    wb = load_workbook(out_path)

    # If sheet_name is empty, use the first sheet
    if not sheet_name:
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[sheet_name]

    header_row, labels = _detect_header_row(ws)
    if not labels:
        raise RuntimeError("Template header not detected.")
    if start_row is None:
        start_row = header_row + 1

    HEADER_JSON_ALIASES: Dict[str, str] = {
        # NEW first column mapping for client name
        "NazevKlienta": "NazevKlienta",
        "Název Klienta": "NazevKlienta",
        "Název klienta*": "NazevKlienta",
        "Prijmeni": "Prijmeni",
        "Příjmení": "Prijmeni",
        "Příjmení*":"Prijmeni",
        "Jmeno": "Jmeno",
        "Jméno": "Jmeno",
        "TitulPred": "TitulPred",
        "TitulZa": "TitulZa",
        "Titul před": "TitulPred",
        "Titul za": "TitulZa",
        "Funkce": "Funkce",
        "Tel1": "Tel1",
        "Tel 1": "Tel1",
        "Telefon": "Tel1",
        "E-mail": "Email",
        "Email": "Email",
        "WWW": "WWW",
        "PoznamkaKOsobe": "PoznamkaKOsobe",
        "Poznámka k osobě": "PoznamkaKOsobe",
    }

    # Build reverse lookup from source rows
    src_keys_norm: Dict[str, str] = {}

    def _normalize_name(s: str) -> str:
        return "".join(ch for ch in str(s).lower() if ch.isalnum())

    for k in set().union(*(r.keys() for r in rows)) if rows else set():
        src_keys_norm[_normalize_name(k)] = k

    label_to_key: Dict[str, str] = {}
    for lbl in labels:
        if lbl in HEADER_JSON_ALIASES:
            label_to_key[lbl] = HEADER_JSON_ALIASES[lbl]
        else:
            # fallback by normalized text
            n = _normalize_name(lbl)
            key = src_keys_norm.get(n)
            if not key and " | " in lbl:
                # try the rightmost token of a multi-row header
                right = lbl.split("|")[-1].strip()
                key = src_keys_norm.get(_normalize_name(right))
                if not key and right in HEADER_JSON_ALIASES:
                    key = HEADER_JSON_ALIASES[right]
            label_to_key[lbl] = key if key else lbl

    aligned = [{lbl: r.get(label_to_key[lbl], "") for lbl in labels} for r in rows]
    df = pd.DataFrame(aligned, columns=labels)

    _unmerge_data_area(ws, first_data_row=start_row)

    r0 = start_row
    for i, (_, s) in enumerate(df.iterrows(), start=r0):
        for j, v in enumerate(s.tolist(), start=1):
            ws.cell(row=i, column=j, value=v)

    # Keep freeze panes if header occupies first two rows and data starts at row 3
    if header_row == 2 and start_row == 3:
        ws.freeze_panes = "A3"

        # --- highlight duplicate (Surname, Name) rows ---
    import unicodedata, re
    def norm(s: str) -> str:
        s = "".join(ch for ch in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(ch))
        return re.sub(r"[^a-z0-9]+", "", s.lower())

    name_col = None
    surname_col = None
    for j, lbl in enumerate(labels, start=1):
        leaf = lbl.split("|")[-1].strip()  # rightmost part like "Jméno" or "Příjmení *"
        n = norm(leaf)
        if n == "jmeno":
            name_col = j
        elif n == "prijmeni":
            surname_col = j

    if name_col and surname_col:
        from collections import defaultdict
        buckets = defaultdict(list)
        row_start = start_row
        row_end = start_row + len(df) - 1
        for r in range(row_start, row_end + 1):
            ln = ws.cell(row=r, column=surname_col).value
            fn = ws.cell(row=r, column=name_col).value
            last_name = (str(ln).strip() if ln is not None else "")
            first_name = (str(fn).strip() if fn is not None else "")
            if last_name or first_name:
                buckets[(last_name.lower(), first_name.lower())].append(r)

        dup_rows = {rr for rr_list in buckets.values() if len(rr_list) > 1 for rr in rr_list}
        if dup_rows:
            # ARGB light red
            fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
            for r in dup_rows:
                for c in range(1, len(labels) + 1):
                    ws.cell(row=r, column=c).fill = fill
    wb.save(out_path)
